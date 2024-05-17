import json
import os
from typing import List, Any

from django.conf.global_settings import STATIC_ROOT
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

# - import all custom forms
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
import pandas as pd

from core.settings import BASE_DIR
from .forms import CreateUserForm, LoginForm, FilesForm
from django.contrib.auth.decorators import login_required

# - Authentication models and functions
from django.contrib.auth.models import auth
from django.contrib.auth import authenticate, login, logout

from pathlib import Path
import openpyxl

from .models import Course


def homepage(request):
    # return HttpResponse("This is the home page.")
    return redirect('login')


def register(request):
    form = CreateUserForm()

    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            form.save()  # pushes all the valid data to the database
            return redirect("login")

    context = {'registerform': form}
    # pushes context data into register form
    return render(request, 'users/register.html', context=context)


def login(request):
    if request.user.is_authenticated:
        return redirect('show')
    form = LoginForm()
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            # check if the database username, password matches
            user = authenticate(request, username=username, password=password)

            if user is not None:
                auth.login(request, user)
                # return redirect("dashboard")
                if user.is_superuser:
                    return redirect('/admin/')  # Redirect admin to the default Django admin page
                else:
                    return redirect('show')  # Redirect faculty to the homepage

    context = {'loginform': form}
    return render(request, 'sign-in.html', context=context)


@login_required(login_url="login.html")
def logout(request):
    auth.logout(request)
    return redirect("login.html")


def upload_excel_files(request):
    if request.method == 'POST':
        BASE_DIR = Path(__file__).resolve().parent.parent
        save_path = os.path.join(os.path.join(BASE_DIR, 'media'), 'ExcelFiles')
        os.makedirs(save_path, exist_ok=True)

        file_storage = FileSystemStorage(location=save_path)

        files = request.FILES.getlist('file')
        if not files:
            print("No files received")
        for file in files:
            print('Files found')
            filename = file_storage.save(file.name, file)

        return redirect("list_files")

    return render(request, 'FileHandling/handleExcelFile.html')


def list_files(request):
    BASE_DIR = Path(__file__).resolve().parent.parent
    save_path = os.path.join(BASE_DIR, 'media', 'ExcelFiles')

    if not os.path.exists(save_path) or not os.path.isdir(save_path):
        return HttpResponse("Directory does not exist or is not accessible", status=404)

    files = []
    for filename in os.listdir(save_path):
        filepath = os.path.join(save_path, filename)
        if os.path.isfile(filepath):
            files.append({
                'name': filename,
                'path': filepath
            })

    return render(request, 'FileHandling/filesList.html', {'files': files})


def delete_file(request):
    try:
        data = json.loads(request.body)
        file_name = data.get('filename')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)

    BASE_DIR = Path(__file__).resolve().parent.parent
    file_path = os.path.join(BASE_DIR, 'media', 'ExcelFiles', file_name)

    if not os.path.exists(file_path):
        return JsonResponse({'success': False, 'error': 'File not found'}, status=404)

    try:
        os.remove(file_path)
        return JsonResponse({'success': True, 'message': 'File deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url="login")
def dashboard(request):
    return render(request, 'users/dashboard.html')


def get_average_mark_from_file(filepath, sheet_name):
    workbook = load_workbook(filename=filepath, data_only=True)
    worksheet = workbook[sheet_name]
    cell_value = worksheet['CZ66'].value
    workbook.close()
    return cell_value


# @login_required(login_url="login")
def dashboard(request):
    BASE_DIR = Path(__file__).resolve().parent.parent
    excel_files_path = os.path.join(BASE_DIR, 'media', 'ExcelFiles')
    courses = {}

    for filename in os.listdir(excel_files_path):
        if filename.endswith('.xlsx'):
            course_code = filename.rsplit(' - ', 1)[0]
            year = filename.rsplit(' - ', 1)[1].split('.')[0].split(' ')[
                len(filename.rsplit(' - ', 1)[1].split('.')[0].split(' ')) - 1]
            filepath = os.path.join(excel_files_path, filename)

            try:
                excel_file = pd.ExcelFile(filepath)
                sheet_name = excel_file.sheet_names[1]
                average_marks = get_average_mark_from_file(filepath, sheet_name)
                if average_marks is None:
                    print(f"The cell CZ66 is empty in the file {filename}")
                    continue

                if average_marks < 1:
                    average_marks = average_marks * 100

                if course_code not in courses:
                    courses[course_code] = []
                courses[course_code].append({'year': year, 'value': average_marks})
            except Exception as e:
                print(f"Error reading {filename}: {e}")
                continue

    chart_data = {
        'labels': sorted(set(year for course in courses.values() for data in course for year in [data['year']])),
        'datasets': []
    }

    for course, data in courses.items():
        sorted_data = sorted(data, key=lambda x: x['year'])
        chart_data['datasets'].append({
            'label': course,
            'data': [d['value'] for d in sorted_data],
            'fill': 'false',
        })

    # return render(request, 'users/index.html', {'chart_data': chart_data})
    render(request, 'uploadeFile.html')


def show(request):
    if not request.user.is_authenticated:
        return redirect('login')
    uuid = request.GET.get('id')
    course = None
    if uuid:
        course = Course.objects.get(id=uuid)
        workbook = openpyxl.load_workbook(os.path.join(BASE_DIR, 'files/' + course.file.name), data_only=True)
    else:
        workbook = openpyxl.load_workbook(os.path.join(BASE_DIR, 'files/null6.xlsx'), data_only=True)
    worksheet = workbook.worksheets[1]
    cellCountStudent = worksheet['Q5'].value

    cell_range = worksheet['D153:L153']
    range_data: list[list[Any]] = [[cell.value for cell in row] for row in cell_range]
    print(range_data[0])
    data = {
        'course': {
            'name': course.name if course is not None else 'default course data ',
            'year': course.year if course is not None else '2024',
            'semester': course.semester if course is not None else '1'
        },
        'courses': Course.objects.get(name=course.name) if course is not None else [],
        'countStudent': cellCountStudent,
        'countA': range_data[0][1],
        'countB': range_data[0][3],
        'countC': range_data[0][5],
        'countA1': range_data[0][0],
        'countB1': range_data[0][2],

    }

    return render(request, 'showChart.html', data)


def api(request):
    c = Course.objects.filter(name=request.GET.get('name'), year=request.GET.get('year'),
                              semester=request.GET.get('semester'))
    # print(c.name)

    if c:
        workbook = openpyxl.load_workbook(os.path.join(BASE_DIR, 'files/' + c[0].file.name), data_only=True)
        worksheet = workbook.worksheets[1]

        cell_range = worksheet['D153:L153']
        range_data: list[list[Any]] = [[cell.value for cell in row] for row in cell_range]

        performanc1: list[list[Any]] = [[cell.value for cell in row] for row in worksheet['E147:M147']]
        performanc2: list[list[Any]] = [[cell.value for cell in row] for row in worksheet['O147:W147']]
        performanc3: list[list[Any]] = [[cell.value for cell in row] for row in worksheet['Y147:AG147']]
        performanc4: list[list[Any]] = [[cell.value for cell in row] for row in worksheet['AI147:AQ147']]
        performanc5: list[list[Any]] = [[cell.value for cell in row] for row in worksheet['AS147:BA147']]
        cellCountStudent = worksheet['Q5'].value

        # values = {key: value for key, value in zip(range_keys[0], range_data[0])}
        data = {
            'countA': range_data[0][1],
            'countA1': range_data[0][0],
            'countB': range_data[0][3],
            'countC': range_data[0][5],
            'countB1': range_data[0][2],
            'data': range_data[0],
            'RateMarkers': worksheet['CZ145'].value,
            'Instructor': worksheet['B158'].value,
            'assessment': [
                {'name': worksheet['D9'].value, 'performanc': performanc1[0], 'rate': worksheet['D145'].value,
                 'lo': [worksheet['F143'].value / (cellCountStudent * worksheet['F12'].value),
                        worksheet['G143'].value / (cellCountStudent * worksheet['G12'].value),
                        (worksheet['E143'].value) / (cellCountStudent * worksheet['E12'].value)]},
                {'name': worksheet['N9'].value, 'performanc': performanc2[0], 'rate': worksheet['N145'].value,
                 'lo': [(worksheet['N143'].value) / (cellCountStudent * worksheet['N12'].value),
                        (worksheet['O143'].value) / (cellCountStudent * worksheet['O12'].value),
                        (worksheet['P143'].value) / (cellCountStudent * worksheet['P12'].value)]},
                {'name': worksheet['X9'].value, 'performanc': performanc3[0], 'rate': worksheet['X145'].value,
                 'lo': [(worksheet['X143'].value) / (cellCountStudent * worksheet['X12'].value),
                        (worksheet['Y143'].value) / (cellCountStudent * worksheet['Y12'].value),
                        (worksheet['Z143'].value) / (cellCountStudent * worksheet['Z12'].value)]},
                {'name': worksheet['AH9'].value, 'performanc': performanc4[0], 'rate': worksheet['AH145'].value,
                 'lo': [(worksheet['AH143'].value) / (cellCountStudent * worksheet['AH12'].value),
                        (worksheet['AI143'].value) / (cellCountStudent * worksheet['AI12'].value),
                        (worksheet['AJ143'].value) / (cellCountStudent * worksheet['AJ12'].value)]},
                {'name': worksheet['AR9'].value, 'performanc': performanc5[0], 'rate': worksheet['AR145'].value,
                 'lo': [(worksheet['AT143'].value) / (cellCountStudent * worksheet['AT12'].value),
                        (worksheet['AS143'].value) / (cellCountStudent * worksheet['AS12'].value),
                        (worksheet['AU143'].value) / (cellCountStudent * worksheet['AU12'].value)]},
            ]

        }
    else:
        data = None
    print(data)
    return JsonResponse(data)


def createFile(request):
    if not request.user.is_authenticated:
        return redirect('login')
    c = Course.objects.all()
    return render(request, 'uploadeFile.html', {'courses': c})


allowed_extensions = ['xlsx']


def storeFile(request):
    print('start')
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == 'POST':
        print('1')
        form = FilesForm(request.POST, request.FILES)
        if form.is_valid():
            print('2')
            files = form.cleaned_data["file"]
            for filename in files:
                file_extension = filename.name.split('.')[-1].lower()
                if file_extension in allowed_extensions:
                    pdf = Course.objects.create(name=request.POST.get('name'),
                                                year=request.POST.get('year'),
                                                semester=request.POST.get('semester'),
                                                user_id=request.POST.get('user_id'),
                                                file=filename)

                    pdf.save()
                    print('3')

            return redirect('create')
        else:
            return redirect('create')
    else:
        return redirect('create')
    # return  HttpResponse('asdfasfd')
    # return render(request, 'upload_files.html')


def profile(request):
    return render(request, 'profile.html')


def deleteExel(request):
    if request.GET.get('uuid'):
        pdf = get_object_or_404(Course, pk=request.GET.get('uuid'))
        if os.path.exists(pdf.file.path):
            os.remove(pdf.file.path)
        pdf.delete()
        return redirect('create')
    else:
        return redirect('create')
